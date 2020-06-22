"""
Script designed to process the A6 Estimate Excel file finding the 
columns needed to extract appropriate data and extracting individual
row cost codes into an array of JSON objects. Script prints one out of
three possible outputs:
- Master error if the file extension is not supported, if required worksheet
  is missing or if the required columns are missing, e.g.:
  [{"ERROR": 'You have selected an invalid file. The estimate file must be of type .xlsx or .xlsm'}]
- List of data errors if any of the processed rows contain invalid data, e.g:
  [{"DATA": "INVALID"}, {"FIELD": "CODE", "LOCATION": "G14"}, ...]
- List of valid cost codes if there is no data errors in the file, e.g.:
  [{"DATA": "VALID"}, {"CODE": "",
                       "COST TYPE": "",
                       "PHASE": "",
                       "LOCATION": "",
                       "DESCRIPTION": "",
                       "QTY.": "",
                       "UNITS": "",
                       "UNIT PRICE": "",
                       "ESTIMATED AMOUNT": "",
                       "GROUPING NAME": "",
                       "SUMMARY NAME": ""}, ...]

Script design and implementation by Michal Zarnowski and Hannah Cheng

"""

from openpyxl import load_workbook, Workbook
import pandas as pd
import numpy as np
import sys
import re
import copy
import json
from string import ascii_uppercase


class CleanUpML:
    sheet = None

    startRowIndex = 0
    rowIndex = 0
    tempHeader = ""
    tempFooter = ""
    tempFooterIndex = 0

    cleanData = [{"DATA": "VALID"}]
    errorData = [{"DATA": "INVALID"}]
    dataTemplate = {
        "CODE": "",
        "COST TYPE": "",
        "PHASE": "",
        "LOCATION": "",
        "DESCRIPTION": "",
        "QTY.": "",
        "UNITS": "",
        "UNIT PRICE": "",
        "ESTIMATED AMOUNT": "",
        "GROUPING NAME": "",
        "SUMMARY NAME": ""
    }

    usableColumns = {
        "MATERIAL UNIT": 0,
        "LABOUR UNIT": 0,
        "LOCATION": 0,
        "PHASE": 0,
        "CODE": 0,
        "DESCRIPTION": 0,
        "QTY": 0,
        "UNITS": 0,
    }

    masterError = {"ERROR": ""}

    # path = 'C:\\Users\\zarnowm\\Documents\\GitHub\\materialLabourConverter\\Template.xlsm'
    path = sys.argv[1]

    def __init__(self):
        pass

    def loadWorkbook(self, path):
        """
        Loads the Excel workbook and extracts the sheet 'Est. Summary'.
        Sheet name has to be the exact match to 'Est. Summary'.

        """

        try:
            wb = load_workbook(filename=path, data_only=True)
            self.sheet = wb['Est. Summary']
        # Handle any possible exceptions resulting from incorrect file format/structure
        except Exception as e:
            if("file format" in str(e)):
                self.masterError['ERROR'] = 'You have selected an invalid file. The estimate file must be of type .xlsx or .xlsm'
            elif("Worksheet" in str(e)):
                self.masterError['ERROR'] = 'The file must have a worksheet titled "Est. Summary". Please ensure that worksheet exists in the selected file'
            else:
                self.masterError['ERROR'] = 'Master error, please contact help for further assitance'

        # If errors occured, print and exit
        if(len(self.masterError['ERROR']) > 0):
            jsonError = json.dumps(self.masterError)
            print(jsonError)
            exit()

    def getHeaderRows(self):
        """
        Scans the file for two header rows of the document.

        """

        maxCol = len(self.sheet.column_dimensions)
        maxRow = len(self.sheet.row_dimensions)

        # Create a map of the excel file in usable format
        text = []
        for value in self.sheet.iter_rows(min_row=1,
                                          max_row=maxRow,
                                          min_col=1,
                                          max_col=maxCol,
                                          values_only=True):
            text.append(value)

        # Scan the map for cell containing "CS" which indicates first cell of the header
        for i in text:
            for j in i:
                if j == 'CS':
                    # File data starts after two headers
                    self.startRowIndex = text.index(i) + 3
                    self.rowIndex = text.index(i) + 3

        # If "CS" wasn't found, return an error and exit, otherwise proceed to find required columns
        if(self.startRowIndex == 0):
            self.masterError['ERROR'] = 'Column Heading error, please ensure the template header structure is unchanged. Missing header: CS CODE'
            jsonError = json.dumps(self.masterError)
            print(jsonError)
            exit()
        else:
            self.findUsableColumns(
                text[self.startRowIndex - 3], text[self.startRowIndex - 2])

    def findUsableColumns(self, rawHeaderOne, rawHeaderTwo):
        """
        Extracts the column position of all columns required for script's functionality.

        """

        # Clean header rows of white spaces
        headerRowOne = self.stripWhiteSpaces(rawHeaderOne)
        headerRowTwo = self.stripWhiteSpaces(rawHeaderTwo)

        # MATERIAL UNIT COLUMN
        # Look for "MAT." in first header row
        if("MAT." in headerRowOne):
            self.usableColumns['MATERIAL UNIT'] = headerRowOne.index("MAT.")

        # LABOUR UNIT COLUMN
        # Look for "LABUNIT" in first header row
        if("LABUNIT" in headerRowOne):
            self.usableColumns['LABOUR UNIT'] = headerRowOne.index("LABUNIT")

        # LOCATION COLUMN
        # Look for "LOCATION" in second header row
        if("LOCATION" in headerRowTwo):
            self.usableColumns['LOCATION'] = headerRowTwo.index(
                "LOCATION")

        # PHASE COLUMN
        # Look for "PHASE" in second header row
        if("PHASE" in headerRowTwo):
            self.usableColumns['PHASE'] = headerRowTwo.index(
                "PHASE")

        # CODE COLUMN
        # Check if "CODE" in cell after "PHASE" cell in second header row
        if(headerRowTwo[self.usableColumns['PHASE'] + 1] == "CODE"):
            self.usableColumns['CODE'] = self.usableColumns['PHASE'] + 1

        # DESCRIPTION COLUMN
        # Look for "DESCRIPTION" in second header row
        if("DESCRIPTION" in headerRowTwo):
            self.usableColumns['DESCRIPTION'] = headerRowTwo.index(
                "DESCRIPTION")

        # QTY COLUMN
        # Look for "QTY." in second header row
        if("QTY." in headerRowTwo):
            self.usableColumns['QTY'] = headerRowTwo.index(
                "QTY.")

        # UNITS COLUMN
        # Check if only one column between "QTY" and "DESCRIPTION"
        if((self.usableColumns['DESCRIPTION'] - self.usableColumns['QTY']) == 2):
            self.usableColumns['UNITS'] = self.usableColumns['QTY'] + 1

        # If any columns were invalid, add to missingColumnList
        missingColumnList = []
        for key, value in self.usableColumns.items():
            if(value == 0):
                missingColumnList.append(key)
        missingColumns = ', '.join(missingColumnList)

        # Print error message with missing columns and exit
        if(len(missingColumnList) > 0):
            self.masterError['ERROR'] = 'Column Heading error, please ensure the template header structure is unchanged. Missing headers: ' + missingColumns
            jsonError = json.dumps(self.masterError)
            print(jsonError)
            exit()

    def stripWhiteSpaces(self, input):
        """
        Strips all white spaces from entities inside supplied list

        """

        stripped = []

        for i in input:
            j = str(i).replace(" ", "")
            stripped.append(j)

        return stripped

    def digestRows(self):
        """
        Iterates over workable rows of the excel sheet and dispatches function calls
        based on contents of each row. Once end of file is reached, class level list
        of create dictionaries "self.cleanData" is converted to JSON format and printed
        to stdout.

        """

        # Iterate over all workable rows
        for row in self.sheet.iter_rows(min_row=self.startRowIndex, values_only=True):

            # Skip empty row
            emptyRow = self.checkIfEmptyRow(row)
            if(emptyRow):
                self.rowIndex += 1
                continue
            # Skip footer and footer preceeding row
            elif(self.rowIndex == self.tempFooterIndex or self.rowIndex == (self.tempFooterIndex - 1)):
                self.rowIndex += 1
                continue
            else:
                if(not self.checkIfHeaderRow(row)):
                    self.createLabourObj(row)
                    self.createMaterialObj(row)
                self.rowIndex += 1

        # Convert output to JSON format and print
        jsonData = json.dumps(self.cleanData)
        jsonErrors = json.dumps(self.errorData)

        if(len(self.errorData) > 1):
            print(jsonErrors)
        else:
            print(jsonData)

        # SAVE OUTPUT IN TXT
        # f = open('output.txt', 'w')
        # print(jsonData, file=f)

    def checkIfEmptyRow(self, row):
        """
        Checks if supplied row is empty. Returns True for empty row
        or False otherwise.

        """
        # Loop through all usable columns to check for cell values.
        for i in range(0, 22):
            if row[i] is not None:
                return False
        return True

    def checkIfHeaderRow(self, row):
        """
        Checks if supplied row is a section header. If true, new temporary
        header is assigned, function to find section footer is called and
        True is returned. If false, False is returned.

        """
        # If value in description column is not null and a upper case String, this is a header row
        if(row[self.usableColumns['DESCRIPTION']] is not None and
           isinstance(row[self.usableColumns['DESCRIPTION']], str) and row[self.usableColumns['DESCRIPTION']].isupper()):
            # Assign section header
            self.tempHeader = row[self.usableColumns['DESCRIPTION']]
            # Find section footer
            self.findSiblingFooter()
            return True
        else:
            return False

    def findSiblingFooter(self):
        """
        Finds the footer of the current section based on current row index. Temporary
        footer and its index are assigned.

        """

        tempRowIndex = self.rowIndex

        # Loop through rows starting on current row
        for row in self.sheet.iter_rows(min_row=self.rowIndex, values_only=True):
            # If value in description column containes minimu of 3 "*", next row is the footer
            if row[self.usableColumns['DESCRIPTION']] is not None and '***' in row[self.usableColumns['DESCRIPTION']]:
                # Assign section footer
                self.tempFooter = self.sheet[tempRowIndex +
                                             1][self.usableColumns['DESCRIPTION']].value
                # Assign section footer index
                self.tempFooterIndex = tempRowIndex + 1
                return
            tempRowIndex += 1

    def createLabourObj(self, row):
        """
        Checks if row contains "Labour" unit price and calls the function to create
        cost code dictionary with appropriate data.

        """

        try:
            # Get labour unit price column index
            labourUnitPriceCol = self.usableColumns['LABOUR UNIT']

            # If labour unit price column contains dashes "-", return without making labour obj
            if(re.search(r'^[\-]*$', str(row[labourUnitPriceCol])) is not None):
                return
            else:
                # Check if labour unit cell contains a valid currency value
                validLabourUnitPrice = self.validateUnitPrice(
                    row[labourUnitPriceCol])

                # If invalid, create error dictionary and add to class' errorData
                if(not validLabourUnitPrice):
                    errorDict = {"FIELD": "LABOUR UNIT PRICE",
                                 "LOCATION": "{}{}".format(ascii_uppercase[labourUnitPriceCol], self.rowIndex)}
                    # Only add if not already in the list (potential duplicate if row contains
                    # labour and material cost code)
                    if(errorDict not in self.errorData):
                        self.errorData.append(errorDict)

                # If value in "LABOUR UNIT" is a valid currency value, call function to create cost code obj
                else:
                    self.convertRowToObj(row, "Labour")
        except:
            pass

    def createMaterialObj(self, row):
        """
        Checks if row contains "Material" unit price and calls the function to create
        cost code dictionary with appropriate data.

        """

        try:
            # Get material unit price column index
            materialUnitPriceCol = self.usableColumns['MATERIAL UNIT']

            # If labour unit price column contains dashes "-", return without making labour obj
            if(re.search(r'^[\-]*$', str(row[materialUnitPriceCol])) is not None):
                return
            else:
                # Check if material unit cell contains a valid currency value
                validMaterialUnitPrice = self.validateUnitPrice(
                    row[materialUnitPriceCol])

                # If invalid, create error dictionary and add to class' errorData
                if(not validMaterialUnitPrice):
                    errorDict = {"FIELD": "MATERIAL UNIT PRICE",
                                 "LOCATION": "{}{}".format(ascii_uppercase[materialUnitPriceCol], self.rowIndex)}
                    # Only add if not already in the list (potential duplicate if row contains
                    # labour and material cost code)
                    if(errorDict not in self.errorData):
                        self.errorData.append(errorDict)

                # If value in "MATERIAL UNIT" is a valid currency value, call function to create cost code obj
                else:
                    self.convertRowToObj(row, "Material")
        except:
            pass

    def convertRowToObj(self, row, objType):
        """
        Converts supplied row to a dictionary and saves it in the class level
        variable "cleanData"

        """

        # Determine column position of unit price based on cost type
        if(objType == "Labour"):
            unitPriceColumn = self.usableColumns['LABOUR UNIT']
        else:
            unitPriceColumn = self.usableColumns['MATERIAL UNIT']

        # VALIDATION
        validRow = self.validateRow(row, unitPriceColumn)
        # If validation failed, return without adding row to cleanData
        if(not validRow):
            return

        try:
            # Create new copy of dictionary template
            newObj = copy.deepcopy(self.dataTemplate)

            # Convert code to 'dd dd dd' format
            rawCode = row[self.usableColumns['CODE']
                          ].replace(" ", "").replace("-", "")  # Change code to [dddddd] format
            code = rawCode[:2] + " " + rawCode[2:-2] + " " + rawCode[-2:]

            # Assign appropriate dictionary values
            newObj["CODE"] = code
            newObj["COST TYPE"] = objType
            newObj["PHASE"] = row[self.usableColumns['PHASE']]
            newObj["LOCATION"] = row[self.usableColumns['LOCATION']]
            newObj["DESCRIPTION"] = row[self.usableColumns['DESCRIPTION']]
            newObj["QTY."] = round(row[self.usableColumns['QTY']], 2)
            newObj["UNITS"] = row[self.usableColumns['UNITS']]
            newObj["UNIT PRICE"] = round(row[unitPriceColumn], 2)
            newObj["ESTIMATED AMOUNT"] = round(
                row[self.usableColumns['QTY']] * row[unitPriceColumn], 2)
            newObj["GROUPING NAME"] = self.tempHeader
            newObj["SUMMARY NAME"] = self.tempFooter

            # Add new dictionary to class list
            self.cleanData.append(newObj)
        except:
            pass

    def validateRow(self, row, unitPriceColumn):
        """
        Applies validation rules to each cell in supplied row. If all validations pass,
        True is returned, False otherwise."

        """

        valid = True

        # Validate COST CODE
        codeColumn = self.usableColumns['CODE']
        validCode = self.validateCode(row[codeColumn])
        # If invalid, create error dictionary and add to class' errorData
        if(not validCode):
            errorDict = {"FIELD": "CODE",
                         "LOCATION": "{}{}".format(ascii_uppercase[codeColumn], self.rowIndex)}
            # Only add if not already in the list (potential duplicate if row contains
            # labour and material cost code)
            if(errorDict not in self.errorData):
                self.errorData.append(errorDict)

            valid = False

        # Validate DESCRIPTION
        descColumn = self.usableColumns['DESCRIPTION']
        validDescription = self.validateDescription(row[descColumn])
        # If invalid, create error dictionary and add to class' errorData
        if(not validDescription):
            errorDict = {"FIELD": "DESCRIPTION",
                         "LOCATION": "{}{}".format(ascii_uppercase[descColumn], self.rowIndex)}
            # Only add if not already in the list (potential duplicate if row contains
            # labour and material cost code)
            if(errorDict not in self.errorData):
                self.errorData.append(errorDict)

            valid = False

        # Validate QUANTITY
        qtyColumn = self.usableColumns['QTY']
        validQty = self.validateQty(row[qtyColumn])
        # If invalid, create error dictionary and add to class' errorData
        if(not validQty):
            errorDict = {"FIELD": "QUANTITY",
                         "LOCATION": "{}{}".format(ascii_uppercase[qtyColumn], self.rowIndex)}
            # Only add if not already in the list (potential duplicate if row contains
            # labour and material cost code)
            if(errorDict not in self.errorData):
                self.errorData.append(errorDict)

            valid = False

        # Return True if all validation passed
        return valid

    def validateCode(self, code):
        """
        Checks if supplied cost code is of valid format. Must consist of 6 digits, in
        one of the following formats:
        12 34 56
        123456
        12-34-56

        """

        if(code is None):
            return False

        validCode = re.search(r'^\d{6}$|^\d{2}( |-)\d{2}( |-)\d{2}$', code)
        if(validCode is not None):
            return True

        return False

    def validateDescription(self, description):
        """
        Checks if supplied description is of valid format. Must not be empty.

        """

        if(description is None):
            return False
        return True

    def validateQty(self, qty):
        """
        Checks if supplied quantity is of valid format. Must be a number greater than
        or equal to zero. Decimals are allowed. May be empty.

        """

        if(qty is None):
            return True

        validQty = re.search(r'^[0-9]\d*(\.\d+)?$', str(qty))
        if(validQty is not None):
            return True

        return False

    def validateUnitPrice(self, unitPrice):
        """
        Checks if supplied unit price is of valid format. Must be a valid dollar amount (eg 100, 100.00).
        Negative numbers and standard currency formatting allowed (eg - $1000.95 or 1,000.95 or -1,000.95
        or (1,000.95) )

        """

        if(unitPrice is None):
            return False

        # Match currency amount (cents optional), optional thousands separators, optional multi-digit fraction, optional brackets surrounding sum
        validUnitPrice = re.search(
            r'^[$|(|($]?[+-]?[$]?[0-9]{1,3}(?:,?[0-9]{3})*(?:\.[0-9]*)?[)]?$', str(unitPrice))
        if(validUnitPrice is not None):
            return True

        return False

    def main(self):

        self.loadWorkbook(path=self.path)
        self.getHeaderRows()
        self.digestRows()


if __name__ == "__main__":
    cleaned = CleanUpML()
    cleaned.main()


# Original template (Sheetname: Est. Summary):
# https://docs.google.com/spreadsheets/d/1O_UsMCW8P7QoImwR0vGSAuEgOBFcJ4ic/edit#gid=1251347588

# Sample that is used to write this script (Sheetname: Est. Summary):
# https://docs.google.com/spreadsheets/d/1E5f4mfj9bmeCjGMPPDgXycc2UdvWZFns/edit#gid=2031079026

# final UI output we want to see in Radix (use Hanna's copy):
# https://docs.google.com/spreadsheets/d/1d9OD6DO6mSgT0D7WTvtj_YewFvB9aYFvi7SWAJmXnM8/edit?ts=5e70e7bf#gid=635371940
