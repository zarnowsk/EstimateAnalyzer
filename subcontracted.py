from openpyxl import load_workbook, Workbook
import pandas as pd
import numpy as np
import sys
import re
import copy
import json
from string import ascii_uppercase


class CleanUpML:
    sheet = None

    startRowIndex = 0
    rowIndex = 0
    tempHeader = ""
    tempFooter = ""
    tempFooterIndex = 0

    cleanData = [{"DATA": "VALID"}]
    errorData = [{"DATA": "INVALID"}]
    dataTemplate = {
        "CODE": "",
        "DESCRIPTION": "",
        "TOTAL": "",
        "SUBTRADE": ""
    }

    usableColumns = {
        "CODE": 0,
        "DESCRIPTION": 0,
        "TOTAL": 0,
        "SUBTRADE": 0,
    }

    masterError = {"ERROR": ""}

    # path = 'C:\\Users\\zarnowm\\Documents\\GitHub\\subtradesConverter\\TestA1.xlsm'
    path = sys.argv[1]

    def __init__(self):
        pass

    def loadWorkbook(self, path):
        """
        Loads the Excel workbook and extracts the sheet 'Subtrades'.
        Sheet name has to be the exact match to 'Subtrades'.

        """

        try:
            wb = load_workbook(filename=path, data_only=True)
            self.sheet = wb['Subtrades']
        # Handle any possible exceptions resulting from incorrect file format/structure
        except Exception as e:
            if("file format" in str(e)):
                self.masterError['ERROR'] = 'You have selected an invalid file. The estimate file must be of type .xlsx or .xlsm'
            elif("Worksheet" in str(e)):
                self.masterError['ERROR'] = 'The file must have a worksheet titled "Subtrades". Please ensure that worksheet exists in the selected file'
            else:
                self.masterError['ERROR'] = 'Master error, please contact help for further assitance'

        # If errors occured, print and exit
        if(len(self.masterError['ERROR']) > 0):
            jsonError = json.dumps(self.masterError)
            print(jsonError)
            exit()

    def getHeaderRows(self):
        """
        Scans the file for two header rows of the document.

        """

        maxCol = len(self.sheet.column_dimensions)
        maxRow = len(self.sheet.row_dimensions)

        # Create a map of the excel file in usable format
        text = []
        for value in self.sheet.iter_rows(min_row=1,
                                          max_row=maxRow,
                                          min_col=1,
                                          max_col=maxCol,
                                          values_only=True):
            text.append(value)

        # Scan the map for cell containing "STATUS" which indicates first cell of the header
        for i in text:
            for j in i:
                if j == 'STATUS':
                    # File data starts after two headers
                    self.startRowIndex = text.index(i) + 2
                    self.rowIndex = text.index(i) + 2

        # If "STATUS" wasn't found, return an error and exit, otherwise proceed to find required columns
        if(self.startRowIndex == 0):
            self.masterError['ERROR'] = 'Column Heading error, please ensure the template header structure is unchanged. Missing header: STATUS'
            jsonError = json.dumps(self.masterError)
            print(jsonError)
            exit()
        else:
            self.findUsableColumns(
                text[self.startRowIndex - 3], text[self.startRowIndex - 2])

    def findUsableColumns(self, rawHeaderOne, rawHeaderTwo):
        """
        Extracts the column position of all columns required for script's functionality.

        """

        # Clean header rows of white spaces
        headerRowOne = self.stripWhiteSpaces(rawHeaderOne)
        headerRowTwo = self.stripWhiteSpaces(rawHeaderTwo)

        # CODE COLUMN
        # Look for "COST" in first header row and "CODE" in second header row
        if("COST" in headerRowOne and "CODE" in headerRowTwo
           and headerRowOne.index("COST") == headerRowTwo.index("CODE")):
            self.usableColumns['CODE'] = headerRowOne.index("COST")

        # DESCRIPTION COLUMN
        # Look for "DESCRIPTION" in second header row
        if("DESCRIPTION" in headerRowTwo):
            self.usableColumns['DESCRIPTION'] = headerRowTwo.index(
                "DESCRIPTION")

        # TOTAL COLUMN
        # Look for "Final Bid" in first header row
        if("FinalBid" in headerRowOne):
            self.usableColumns['TOTAL'] = headerRowOne.index(
                "FinalBid")

        # SUBTRADE COLUMN
        # Look for "SUBTRADE" in second header row
        if("SUBTRADE" in headerRowTwo):
            self.usableColumns['SUBTRADE'] = headerRowTwo.index(
                "SUBTRADE")

        # If any columns were invalid, add to missingColumnList
        missingColumnList = []
        for key, value in self.usableColumns.items():
            if(value == 0):
                missingColumnList.append(key)
        missingColumns = ', '.join(missingColumnList)

        # Print error message with missing columns and exit
        if(len(missingColumnList) > 0):
            self.masterError['ERROR'] = 'Column Heading error, please ensure the template header structure is unchanged. Missing headers: ' + missingColumns
            jsonError = json.dumps(self.masterError)
            print(jsonError)
            exit()

    def stripWhiteSpaces(self, input):
        """
        Strips all white spaces from entities inside supplied list

        """

        stripped = []

        for i in input:
            j = str(i).replace(" ", "")
            stripped.append(j)

        return stripped

    def digestRows(self):
        """
        Iterates over workable rows of the excel sheet and dispatches function calls
        based on contents of each row. Once end of file is reached, class level list
        of created dictionaries "self.cleanData" is converted to JSON format and printed
        to stdout.

        """

        # Iterate over all workable rows
        for row in self.sheet.iter_rows(min_row=self.startRowIndex, values_only=True):

            # Skip empty row
            emptyRow = self.checkIfEmptyRow(row)
            if(emptyRow):
                self.rowIndex += 1
                continue
            else:
                self.createSubtradeObj(row)
                self.rowIndex += 1

        # Convert output to JSON format and print
        jsonData = json.dumps(self.cleanData)
        jsonErrors = json.dumps(self.errorData)

        if(len(self.errorData) > 1):
            print(jsonErrors)
        else:
            print(jsonData)

        # SAVE OUTPUT IN TXT
        # f = open('output.txt', 'w')
        # print(jsonData, file=f)

    def checkIfEmptyRow(self, row):
        """
        Checks if supplied row is empty. Returns True for empty row
        or False otherwise.

        """
        # Loop through all usable columns to check for cell values.
        for i in range(0, 14):
            if row[i] is not None:
                return False
        return True

    def createSubtradeObj(self, row):
        """
        Converts supplied row to a dictionary and saves it in the class level
        variable "cleanData"

        """

        # If "Final Bid Amount" for this row is not a valid dollar value,
        # or zero, return without creating an object
        validTotal = re.search(
            r'^[$|(|($]?[+-]?[$]?[0-9]{1,3}(?:,?[0-9]{3})*(?:\.[0-9]*)?[)]?$', str(row[self.usableColumns["TOTAL"]]))
        if(validTotal is None or row[self.usableColumns["TOTAL"]] == 0):
            return
        else:
            # VALIDATION
            validRow = self.validateRow(row)

            # If validation failed, return without adding row to cleanData
            if(not validRow):
                return

            try:
                # Create new copy of dictionary template
                newObj = copy.deepcopy(self.dataTemplate)

                # Convert code to 'dd dd dd' format
                rawCode = row[self.usableColumns['CODE']
                              ].replace(" ", "").replace("-", "")  # Change code to [dddddd] format
                code = rawCode[:2] + " " + rawCode[2:-2] + " " + rawCode[-2:]

                # Assign appropriate dictionary values
                newObj["CODE"] = code
                newObj["DESCRIPTION"] = row[self.usableColumns['DESCRIPTION']]
                newObj["TOTAL"] = row[self.usableColumns['TOTAL']]
                newObj["SUBTRADE"] = row[self.usableColumns['SUBTRADE']]

                # Add new dictionary to class list
                self.cleanData.append(newObj)
            except:
                pass

    def validateRow(self, row):
        """
        Applies validation rules to each cell in supplied row. If all validations pass,
        True is returned, False otherwise."

        """

        valid = True

        # Validate COST CODE
        codeColumn = self.usableColumns['CODE']
        validCode = self.validateCode(row[codeColumn])
        # If invalid, create error dictionary and add to class' errorData
        if(not validCode):
            errorDict = {"FIELD": "CODE",
                         "LOCATION": "{}{}".format(ascii_uppercase[codeColumn], self.rowIndex)}
            self.errorData.append(errorDict)
            valid = False

        # Validate DESCRIPTION
        descColumn = self.usableColumns['DESCRIPTION']
        validDescription = self.validateDescriptionAndSubtrade(row[descColumn])
        # If invalid, create error dictionary and add to class' errorData
        if(not validDescription):
            errorDict = {"FIELD": "DESCRIPTION",
                         "LOCATION": "{}{}".format(ascii_uppercase[descColumn], self.rowIndex)}
            self.errorData.append(errorDict)
            valid = False

        # Validate SUBTRADE
        subtradeColumn = self.usableColumns['SUBTRADE']
        validSubtrade = self.validateDescriptionAndSubtrade(
            row[subtradeColumn])
        # If invalid, create error dictionary and add to class' errorData
        if(not validSubtrade):
            errorDict = {"FIELD": "SUBTRADE",
                         "LOCATION": "{}{}".format(ascii_uppercase[subtradeColumn], self.rowIndex)}
            self.errorData.append(errorDict)
            valid = False

        # Return valid/invalid boolean
        return valid

    def validateCode(self, code):
        """
        Checks if supplied cost code is of valid format. Must consist of 6 digits, in
        one of the following formats:
        12 34 56
        123456
        12-34-56

        """

        if(code is None):
            return False

        validCode = re.search(r'^\d{6}$|^\d{2}( |-)\d{2}( |-)\d{2}$', code)
        if(validCode is not None):
            return True

        return False

    def validateDescriptionAndSubtrade(self, cellValue):
        """
        Checks if supplied description or subtrade is of valid format. Must not be empty.

        """

        if(cellValue is None):
            return False
        return True

    def main(self):

        self.loadWorkbook(path=self.path)
        self.getHeaderRows()
        self.digestRows()


if __name__ == "__main__":
    cleaned = CleanUpML()
    cleaned.main()
