from openpyxl import load_workbook, Workbook
import pandas as pd
import numpy as np


class CleanUpML:
    sheet = None

    row_index = 0
    col_index = 0

    cleaned_df = None
    header_df = None
    material_only = None
    labour_only = None
    material_and_labour = None
    material_within_ml = None
    labour_within_ml = None
    merged_df = None
    merged_df_with_header = None

    path = '/Users/chengh/Documents/costing/Book1.xlsx'

    def __init__(self):
        pass

    def load_workbook(self, path):
        """
        Loads the Excel workbook and extracting the sheet 'Est. Summary';
        sheet name has to be the exact match to 'Est. Summary'
        **File has to be in .xlsm format
        """

        wb = load_workbook(filename=path)
        self.sheet = wb['Est. Summary']

    def get_index(self):
        """
        Return the Excel index for row and column at 'CS', the first cell the program
        will start reading from

        If we are not allowing the flexibility to modify rows above
        'CS', then this function can be removed
        """

        max_col = len(self.sheet.column_dimensions)
        max_row = len(self.sheet.row_dimensions)

        text = []

        for value in self.sheet.iter_rows(min_row=1,
                                          max_row=max_row,
                                          min_col=1,
                                          max_col=max_col,
                                          values_only=True):
            text.append(value)

        for i in text:
            for j in i:
                if j == 'CS':
                    self.row_index = text.index(i)
                    self.col_index = i.index(j)

        # print(self.row_index, self.col_index)

    def clean_header(self):
        """
        Extracting and renaming relevant headers to the correct names and in order
        """

        df = pd.read_excel(self.path, sheet_name='Est. Summary')

        df = df.iloc[self.row_index - 1:, self.col_index:]

        df.columns = df.iloc[:2, :].fillna('').apply(' '.join).str.strip()

        df.columns.values[3] = 'UNITS'

        df.rename(columns={"D E S C R I P T I O N": "DESCRIPTION",
                           'LABOUR  TOTAL': "LABOUR TOTAL",
                           'ED COST CODE LOCATION': "LOCATION"}, inplace=True)

        df = df[['CODE', 'DESCRIPTION', 'LOCATION', 'PHASE', 'QTY.', 'UNITS',
                 'MAT. UNIT', 'MATERIAL TOTAL', 'LAB UNIT INC P./ B.',
                 'LABOUR TOTAL']]

        self.cleaned_df = df[2:]

        self.header_df = df[2:]

        # print(self.cleaned_df.shape, self.header_df.shape)
        # print(self.cleaned_df.columns, self.header_df.columns)

    def get_last_row(self):
        """
        Return the last row of estimate code to read
        """

        last_row_index = \
            self.cleaned_df.index[self.cleaned_df['DESCRIPTION'] == 'Payroll Burden For Work Above 3rd Flr (Ont Only)'][
                0]
        last_row_number = self.cleaned_df.index.get_loc(last_row_index)
        self.cleaned_df = self.cleaned_df.iloc[:last_row_number + 1, :]
        self.header_df = self.cleaned_df
        # print(last_row_number, last_row_index, self.cleaned_df.shape)

    def clean_material_col(self):
        """
        Replacing string inputs in material column with 0
        ** this method will be removed when incorporating error raising
        """

        string_inputs_material_total = self.cleaned_df[(self.cleaned_df['MATERIAL TOTAL'].str.isdigit() == False)][
            'MATERIAL TOTAL'].unique()

        for i in string_inputs_material_total:
            self.cleaned_df['MATERIAL TOTAL'].replace(i, 0, inplace=True)

        # print(self.cleaned_df.shape)

    def clean_labour_col(self):
        """
        Replacing string inputs in labour column with 0
        ** this method will be removed when incorporating error raising
        """

        string_inputs_labour_total = self.cleaned_df[(self.cleaned_df['LABOUR TOTAL'].str.isdigit() == False)][
            'LABOUR TOTAL'].unique()

        for i in string_inputs_labour_total:
            self.cleaned_df['LABOUR TOTAL'].replace(i, 0, inplace=True)

        # print(self.cleaned_df.shape)
        # print(self.cleaned_df)

    def material_only(self):
        """
        Extract material only estimate code: codes that are 0 or nan in labour total,
        and not nan and not 0 in material total
        """

        self.material_only = self.cleaned_df[
            (self.cleaned_df['LABOUR TOTAL'] == 0 | self.cleaned_df['LABOUR TOTAL'].isna()) &
            (self.cleaned_df['MATERIAL TOTAL'].notna() & self.cleaned_df['MATERIAL TOTAL'] != 0)]

        self.material_only.insert(self.material_only.shape[1], 'COST TYPE', 'Material')

        self.material_only = self.material_only.rename(columns={'MATERIAL TOTAL': 'ESTIMATED AMOUNT',
                                                                'MAT. UNIT': 'UNIT PRICE'})

        self.material_only.drop(columns=['LAB UNIT INC P./ B.', 'LABOUR TOTAL'], inplace=True)

        # print("material only: ", self.material_only.shape)
        # print(self.material_only.columns)
        # print(self.material_only)

    def labour_only(self):
        """
        Extract labour only estimate code: codes that are 0 or nan in material total,
        and not nan and not 0 in labour total
        """

        self.labour_only = self.cleaned_df[
            (self.cleaned_df['MATERIAL TOTAL'] == 0 | self.cleaned_df['MATERIAL TOTAL'].isna()) &
            (self.cleaned_df['LABOUR TOTAL'].notna() & self.cleaned_df['LABOUR TOTAL'] != 0)]

        self.labour_only.insert(self.labour_only.shape[1], 'COST TYPE', 'Labour')

        self.labour_only = self.labour_only.rename(columns={'LABOUR TOTAL': 'ESTIMATED AMOUNT',
                                                            'LAB UNIT INC P./ B.': 'UNIT PRICE'})
        #
        self.labour_only.drop(columns=['MAT. UNIT', 'MATERIAL TOTAL'], inplace=True)

        # print("labour only: ", self.labour_only.shape)

    def material_and_labour(self):
        """
        Extract material and labour estimate code: codes that are 0 or nan in labour total,
        and not nan and not 0 in material total
        """

        self.material_and_labour = self.cleaned_df[
            self.cleaned_df['LABOUR TOTAL'].notna() & self.cleaned_df['MATERIAL TOTAL'].notna()
            & (self.cleaned_df['MATERIAL TOTAL'] != 0) & (self.cleaned_df['LABOUR TOTAL'] != 0)]

        # print(self.material_and_labour.shape)

    def material_within_ml(self):
        """
        Extract material estimate from codes that are both labour and material
        """

        self.material_within_ml = self.material_and_labour[['CODE', 'DESCRIPTION', 'LOCATION', 'PHASE', 'QTY.', 'UNITS',
                                                            'MAT. UNIT', 'MATERIAL TOTAL']]

        self.material_within_ml.insert(self.material_within_ml.shape[1], 'COST TYPE', 'Material')

        self.material_within_ml = self.material_within_ml.rename(columns={'MATERIAL TOTAL': 'ESTIMATED AMOUNT',
                                                                          'MAT. UNIT': 'UNIT PRICE'})

        # print("mlm: ", self.material_within_ml.shape)

    def labour_within_ml(self):
        """
        Extract labour estimate from codes that are both labour and material
        """

        self.labour_within_ml = self.material_and_labour[['CODE', 'DESCRIPTION', 'LOCATION', 'PHASE', 'QTY.', 'UNITS',
                                                          'LAB UNIT INC P./ B.', 'LABOUR TOTAL']]

        self.labour_within_ml.insert(self.labour_within_ml.shape[1], 'COST TYPE', 'Labour')

        self.labour_within_ml = self.labour_within_ml.rename(columns={'LABOUR TOTAL': 'ESTIMATED AMOUNT',
                                                                      'LAB UNIT INC P./ B.': 'UNIT PRICE'})

        # print("mll:", self.labour_within_ml.shape)

    def merged_df(self):
        """
        Merging following 4 types of estimate lines into one table:
        1. material only estimates
        2. labour only estimates
        3. material estimates that are both material and labour
        4. labour estimates that are both material and labour
        """

        self.merged_df = pd.concat([self.material_within_ml,
                                    self.labour_within_ml,
                                    self.material_only,
                                    self.labour_only])

        self.merged_df = self.merged_df.sort_index()
        self.merged_df = self.merged_df[['CODE', 'COST TYPE', 'PHASE', 'LOCATION', 'DESCRIPTION', 'QTY.', 'UNITS',
                                         'UNIT PRICE', 'ESTIMATED AMOUNT']]

        # print(self.merged_df.shape)

    def grab_headers(self):
        """
        Extracting header, second header, and summary line for each estimate line

        Condition for header lines: capitalized cells in 'DESCRIPTION' column
        Condition for second header: where only the 'DESCRIPTION' column is not null
        Condition for summary lines: one line below cells that are '**********'
        """

        self.header_df = self.header_df[['CODE', 'DESCRIPTION', 'LOCATION', 'PHASE', 'QTY.', 'UNITS',
                                         'MAT. UNIT', 'MATERIAL TOTAL', 'LAB UNIT INC P./ B.',
                                         'LABOUR TOTAL']]

        self.header_df = self.header_df[~self.header_df['DESCRIPTION'].str.isupper().isna()]

        self.header_df.loc[self.header_df['DESCRIPTION'].str.isupper() == True, 'Level'] = 'header'

        for i in (self.header_df[self.header_df['DESCRIPTION'] == '**********'].index + 1):
            self.header_df.at[i, 'Level'] = 'summary'

        self.header_df.loc[
            (self.header_df.isnull().sum(axis=1) == len(self.header_df.columns) - 1), 'Level'] = 'second header'
        #
        # print(self.header_df.shape)
        # print(self.header_df)

    def merged_df_with_header(self):

        """
        Merging header, second header, summary lines with all the estimate lines
        Removing and rows with the description is '**********'
        """

        header = self.header_df[['DESCRIPTION', 'Level']]

        self.merged_df_with_header = header.join(self.merged_df, rsuffix='_header')

        self.merged_df_with_header['SUBCONTRACT NAME'] = np.nan

        self.merged_df_with_header = self.merged_df_with_header[['CODE', 'COST TYPE', 'PHASE', 'LOCATION',
                                                                 'DESCRIPTION', 'QTY.', 'UNITS', 'UNIT PRICE',
                                                                 'ESTIMATED AMOUNT', 'SUBCONTRACT NAME',
                                                                 'Level']]

        self.merged_df_with_header = self.merged_df_with_header[
            self.merged_df_with_header['DESCRIPTION'] != '**********']

    def insert_grouping_name_col(self):

        """
        This function inserts a column of 'header' that each estimate line belongs to

        This function is for developers who want to group estimate lines by their respective headers

        Doesn't have grouping for the second header

        """

        grouping_name = self.merged_df_with_header[self.merged_df_with_header['Level'] == 'header']
        grouping_name['Grouping Name'] = grouping_name['DESCRIPTION']
        grouping_name = grouping_name['Grouping Name']

        self.merged_df_with_header = self.merged_df_with_header.join(grouping_name, rsuffix='_groupingname')

        self.merged_df_with_header['Grouping Name'].fillna(method='ffill', inplace=True)

        # print(self.merged_df_with_header)

    def insert_summary_name_col(self):

        """
          This function inserts a column of 'summary' that each estimate line belongs to

          This function is for developers who want to group estimate lines by their respective summary lines

          Doesn't have grouping for the second header

        """

        summary_name = self.merged_df_with_header[self.merged_df_with_header['Level'] == 'summary']
        summary_name['Summary Name'] = summary_name['DESCRIPTION']
        summary_name = summary_name['Summary Name']

        self.merged_df_with_header = self.merged_df_with_header.join(summary_name, rsuffix='_summaryname')

        self.merged_df_with_header['Summary Name'].fillna(method='bfill', inplace=True)

        # print(self.merged_df_with_header.shape)
        # print(self.merged_df_with_header.to_json(orient='records'))

    def main(self):

        self.load_workbook(path=self.path)
        self.get_index()
        self.clean_header()
        self.get_last_row()
        self.clean_material_col()
        self.clean_labour_col()
        self.material_only()
        self.labour_only()
        self.material_and_labour()
        self.material_within_ml()
        self.labour_within_ml()
        self.merged_df()
        self.grab_headers()
        self.merged_df_with_header()
        self.insert_grouping_name_col()
        self.insert_summary_name_col()


if __name__ == "__main__":
    cleaned = CleanUpML()
    cleaned.main()
    
    
# Original template (Sheetname: Est. Summary): 
# https://docs.google.com/spreadsheets/d/1O_UsMCW8P7QoImwR0vGSAuEgOBFcJ4ic/edit#gid=1251347588

# Sample that is used to write this script (Sheetname: Est. Summary): 
# https://docs.google.com/spreadsheets/d/1E5f4mfj9bmeCjGMPPDgXycc2UdvWZFns/edit#gid=2031079026

# final UI output we want to see in Radix (use Hanna's copy): 
# https://docs.google.com/spreadsheets/d/1d9OD6DO6mSgT0D7WTvtj_YewFvB9aYFvi7SWAJmXnM8/edit?ts=5e70e7bf#gid=635371940
